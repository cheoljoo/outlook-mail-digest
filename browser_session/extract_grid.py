# -*- coding: utf-8 -*-
"""
Excel Online은 화면에 보이는 셀만 캔버스로 그리기 때문에 (가상 스크롤 + canvas 렌더링),
HTML을 그대로 읽어서는 셀 값을 알 수 없습니다. 클립보드 복사(Ctrl+C 후 clipboard API로
읽기)도 시도해봤지만 브라우저 보안 정책상 권한이 거부됩니다.

대신 셀에 포커스를 옮기면 수식 입력줄(Formula bar)에 그 셀의 실제 값이 텍스트로 나타나는
것을 이용합니다. Ctrl+End로 사용 범위를 먼저 알아낸 뒤, 행마다 이름 상자(Name Box)로
그 행의 첫 칸으로 이동하고 Tab으로 오른쪽으로 이동하며 수식 입력줄을 읽어 openpyxl로
재구성합니다. 셀 수가 많으면(수백~수천) 시간이 꽤 걸립니다 (몇 분 단위).

사용법:
  .venv/bin/python extract_grid.py "<공유 URL>" <저장할 xlsx 경로>
"""

import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook
from playwright.sync_api import sync_playwright

PROFILE_DIR = Path.home() / ".secure_browser_profile"
COL_RE = re.compile(r"^([A-Z]+)(\d+)$")


def col_to_index(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


def set_name_box(wac, ref: str) -> None:
    name_box = wac.get_by_role("combobox", name="이름 상자")
    name_box.click()
    name_box.fill(ref)
    name_box.press("Enter")


def main() -> None:
    if len(sys.argv) < 3:
        print("사용법: extract_grid.py <공유 URL> <저장할 xlsx 경로>")
        sys.exit(1)

    url = sys.argv[1]
    out_path = Path(sys.argv[2])

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(PROFILE_DIR),
            headless=True,
            viewport={"width": 1600, "height": 900},
        )
        page = context.new_page()
        page.goto(url, wait_until="load", timeout=60000)

        wac = None
        for _ in range(30):
            wac = next((fr for fr in page.frames if fr.name == "WacFrame_Excel_0"), None)
            if wac is not None:
                try:
                    wac.get_by_role("combobox", name="이름 상자").wait_for(timeout=2000)
                    break
                except Exception:
                    wac = None
            page.wait_for_timeout(1000)
        if wac is None:
            print("[ERROR] Excel 편집 화면(WacFrame_Excel_0)이 뜨지 않았습니다.")
            sys.exit(1)

        formula_bar = wac.get_by_role("textbox", name="formula bar")
        name_box = wac.get_by_role("combobox", name="이름 상자")

        # 워크북 데이터 모델이 완전히 준비되기 전에 Ctrl+End를 누르면 실제보다 훨씬
        # 작은 범위(예: 화면에 막 보이는 몇 줄)로 잘못 응답하는 경우가 있어서,
        # 연속 두 번 같은 값이 나올 때까지 반복해서 안정화를 확인한다.
        page.wait_for_timeout(3000)
        last_ref = None
        for attempt in range(10):
            set_name_box(wac, "A1")
            page.wait_for_timeout(300)
            wac.locator("body").press("Control+End")
            page.wait_for_timeout(700)
            candidate = name_box.input_value().strip()
            print(f"[INFO] Ctrl+End 시도 {attempt + 1}: {candidate}", flush=True)
            if candidate == last_ref:
                break
            last_ref = candidate
        m = COL_RE.match(last_ref.strip())
        if not m:
            print(f"[ERROR] 마지막 셀 참조를 이해할 수 없습니다: {last_ref!r}")
            sys.exit(1)
        last_col, last_row = m.group(1), int(m.group(2))
        last_col_idx = col_to_index(last_col)
        print(f"[INFO] 사용 범위: A1:{last_ref}  ({last_row}행 x {last_col_idx}열)", flush=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        def read_row(row: int, settle_ms: int, tab_ms: int) -> list:
            set_name_box(wac, f"A{row}")
            page.wait_for_timeout(settle_ms)
            values = []
            for col_idx in range(1, last_col_idx + 1):
                values.append(formula_bar.inner_text())
                if col_idx < last_col_idx:
                    wac.locator("body").press("Tab")
                    page.wait_for_timeout(tab_ms)
            return values

        def write_row(row: int, values: list) -> None:
            for col_idx, value in enumerate(values, start=1):
                if value:
                    ws.cell(row=row, column=col_idx, value=value)

        start = time.time()
        for row in range(1, last_row + 1):
            write_row(row, read_row(row, settle_ms=70, tab_ms=60))

            if row % 20 == 0 or row == last_row:
                elapsed = time.time() - start
                print(f"[INFO] {row}/{last_row}행 완료 ({elapsed:.0f}초 경과)", flush=True)
                wb.save(str(out_path))  # 중간 저장 (중단돼도 여기까지는 남도록)

        # 검증/복구: A열(받은시각)이 타임스탬프 형식이 아니면 그 행이 이전 행의
        # 마지막 칸 값을 잘못 이어받은(stale read) 것일 가능성이 높으므로, 더 긴
        # 대기시간으로 다시 읽어서 덮어쓴다.
        ts_re = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}")
        suspicious = [
            row
            for row in range(2, last_row + 1)
            if not ts_re.match(str(ws.cell(row=row, column=1).value or ""))
        ]
        print(f"[INFO] 검증: 의심스러운 행 {len(suspicious)}개, 재확인합니다.", flush=True)
        for row in suspicious:
            write_row(row, read_row(row, settle_ms=400, tab_ms=250))
        if suspicious:
            wb.save(str(out_path))

        still_bad = [
            row
            for row in range(2, last_row + 1)
            if not ts_re.match(str(ws.cell(row=row, column=1).value or ""))
        ]
        if still_bad:
            print(f"[WARN] 재확인 후에도 이상한 행: {still_bad}", flush=True)
        else:
            print("[INFO] 검증 통과: 모든 행의 받은시각이 정상입니다.", flush=True)

        wb.save(str(out_path))
        print(f"[OK] 저장 완료: {out_path}", flush=True)

        context.close()


if __name__ == "__main__":
    main()
